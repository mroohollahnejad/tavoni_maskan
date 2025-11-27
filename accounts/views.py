from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.contrib.auth import login,logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.http import JsonResponse,HttpResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from .models import Profile, Payment, ApprovedPaymentDate
from .forms import ProfileForm, UserUpdateForm, PaymentForm
import jdatetime
from datetime import datetime
from datetime import date
import pandas as pd
import openpyxl
from .forms import MembersUploadForm
from openpyxl import load_workbook
from datetime import date
import jdatetime
import logging

logger = logging.getLogger(__name__)


# ---------------------- تبدیل تاریخ شمسی یا میلادی به میلادی ----------------------
def parse_date(date_str):
    if not date_str:
        return None
    try:
        jdate = jdatetime.datetime.strptime(date_str, "%Y-%m-%d")
        return jdate.togregorian().date()
    except Exception:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            raise ValueError("فرمت تاریخ معتبر نیست (باید شمسی یا میلادی باشد).")

# ---------------------- تبدیل میلادی به شمسی برای نمایش ----------------------
def to_jalali(date):
    if not date:
        return ''
    if isinstance(date, jdatetime.date):
        return date.strftime("%Y-%m-%d")
    return jdatetime.date.fromgregorian(date=date).strftime("%Y-%m-%d")

# ---------------------- محاسبه امتیاز کل و داده‌های واریزی ----------------------
def calculate_total_score(user):
    payments = Payment.objects.filter(user=user).order_by('installment_number')
    total_score = 0
    payments_data = []
    today = date.today()  # تاریخ امروز
    
    for p in payments:
        if p.payment_date and p.due_date:
            if p.payment_date > p.due_date:
                diff_days = (today - p.payment_date).days
            else:
                diff_days = (today - p.due_date).days
        else:
            diff_days = 0
        
        score = (diff_days * int(p.amount)) / 100_000_000
        total_score += score

        payments_data.append({
            'id': p.id,
            'installment_number': p.installment_number,
            'amount': f"{int(p.amount):,}",  # ← جداکننده هزارگان
            'payment_date': p.payment_date,
            'payment_date_j': to_jalali(p.payment_date),
            'due_date': p.due_date,
            'due_date_j': to_jalali(p.due_date),
            'diff_days': diff_days,
            'score': round(score, 2)
        })
    return total_score, payments_data

# ---------------------- داشبورد ----------------------
@login_required
def dashboard(request):
    user = request.user
    profile = Profile.objects.get_or_create(user=user)[0]
    payments = Payment.objects.filter(user=user).order_by('installment_number')
    approved_dates = ApprovedPaymentDate.objects.all()

    total_score, payments_data = calculate_total_score(user)
    approved_dates_list = {p.installment_number: p.due_date for p in approved_dates}

    # ----------------- محاسبه رتبه -----------------
    users = User.objects.all()
    scores = []
    for u in users:
        score, _ = calculate_total_score(u)
        scores.append((u.id, score))
    scores.sort(key=lambda x: x[1], reverse=True)

    user_rank = None
    rank = 1
    for uid, score in scores:
        if uid == user.id:
            user_rank = rank
            break
        rank += 1

    context = {
        'user': user,
        'profile': profile,
        'payments': payments_data,
        'approved_dates': approved_dates_list,
        'installments': [d.installment_number for d in ApprovedPaymentDate.objects.all().order_by('installment_number')],
        'active_tab': request.GET.get('tab', 'personal'),
        'total_score': total_score,
        "rank": user_rank
    }
    return render(request, 'accounts/dashboard.html', context)

# ---------------------- AJAX ثبت واریزی ----------------------
# ثبت واریزی
@csrf_exempt
@login_required
def payment_create_ajax(request):
    if request.method == 'POST':
        try:
            installment = int(request.POST.get('installment_number'))
            amount = int(float(request.POST.get('amount', 0)))  # ← تبدیل به int
            payment_date = parse_date(request.POST.get('payment_date'))
            due_date_obj = ApprovedPaymentDate.objects.filter(installment_number=installment).first()
            due_date_val = due_date_obj.due_date if due_date_obj else None

            payment = Payment.objects.create(
                user=request.user,
                installment_number=installment,
                amount=amount,
                payment_date=payment_date,
                due_date=due_date_val
            )

            total_score, _ = calculate_total_score(request.user)

            return JsonResponse({
                'status': 'success',
                'message': '✅ واریزی ثبت شد',
                'payment': {
                    'id': payment.id,
                    'installment_number': payment.installment_number,
                    'amount': f"{payment.amount:,}",  # ← جداکننده هزارگان
                    'payment_date': to_jalali(payment.payment_date),
                    'due_date': to_jalali(payment.due_date)
                },
                'total_score': total_score
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر ❌'})

# ویرایش واریزی
@csrf_exempt
@login_required
def payment_edit_ajax(request, pk):
    if request.method == 'POST':
        try:
            payment = get_object_or_404(Payment, pk=pk, user=request.user)
            field = request.POST.get('field')
            value = request.POST.get('value')

            if field == 'installment_number':
                payment.installment_number = int(value)
                due_date_obj = ApprovedPaymentDate.objects.filter(installment_number=value).first()
                payment.due_date = due_date_obj.due_date if due_date_obj else None
            elif field == 'amount':
                payment.amount = int(float(value))  # ← همیشه int
            elif field == 'payment_date':
                payment.payment_date = parse_date(value)

            payment.save()
            total_score, _ = calculate_total_score(request.user)

            return JsonResponse({
                'status': 'success',
                'message': '✅ ویرایش انجام شد',
                'amount': f"{payment.amount:,}",  # ← جداکننده هزارگان
                'payment_date': to_jalali(payment.payment_date),
                'due_date': to_jalali(payment.due_date),
                'total_score': total_score
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر ❌'})

# ---------------------- AJAX حذف واریزی ----------------------
@csrf_exempt
@login_required
def payment_delete_ajax(request, pk):
    if request.method == 'POST':
        payment = get_object_or_404(Payment, pk=pk, user=request.user)
        payment.delete()
        total_score, _ = calculate_total_score(request.user)
        return JsonResponse({'status': 'success', 'message': '🗑️ واریزی حذف شد', 'total_score': total_score})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر ❌'})

# ---------------------- AJAX داده‌های تب امتیاز ----------------------
@login_required
def score_ajax(request):
    total_score, payments_data = calculate_total_score(request.user)
    payments_json = [
        {
            'installment_number': p['installment_number'],
            'amount': int(p.amount),
            'payment_date': p['payment_date_j'],
            'due_date': p['due_date_j'],
            'diff_days': p['diff_days'],
            'score': round(p['score'], 2)
        } for p in payments_data
    ]
    return JsonResponse({'total_score': total_score, 'payments': payments_json})

# ---------------------- ویرایش پروفایل ----------------------
@login_required
def profile_edit(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, "✅ تغییرات ذخیره شد")
            return redirect('dashboard')
        else:
            messages.error(request, "❌ خطا در فرم")
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileForm(instance=profile)
    return render(request, 'accounts/profile_edit.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

# ---------------------- ثبت‌نام کاربر ----------------------
def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        if password != password2:
            messages.error(request, 'رمزها مطابقت ندارند ❌')
            return redirect('register')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'این نام کاربری قبلاً ثبت شده است ❌')
            return redirect('register')
        user = User.objects.create_user(username=username, password=password)
        login(request, user)
        messages.success(request, 'ثبت‌نام با موفقیت انجام شد ✅')
        return redirect('dashboard')
    return render(request, 'accounts/register.html')

# ---------------------------------------------------------------------------



def is_admin(u):
    return u.is_staff or u.is_superuser



# ---------- تولید فایل نمونه ----------
def generate_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = ["نام", "نام خانوادگی", "کد ملی", "شماره همراه", "ایمیل",
               "شماره شناسنامه", "نوبت واریزی", "تاریخ واریز",
               "مبلغ واریزی", "تاریخ مصوب"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    sample_row = ["محمد", "روح‌الله‌نژاد", "۳۳۹۲۰۴۸۴۴۴","09121234567", "test@example.com",
                  "123456", 1, "1404-01-15", 1000000, "1404-01-20"]
    ws.append(sample_row)

    return wb
#------------------------------------------------------------------------------
@login_required
def download_sample_excel(request):
    wb = generate_sample_excel()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_members.xlsx"'
    wb.save(response)
    return response

#-------------------------------------------------------------------------------------
# ---------------------- تبدیل تاریخ شمسی یا میلادی به میلادی ----------------------
def parse_date(date_str):
    """تبدیل تاریخ شمسی یا میلادی به datetime.date"""
    if not date_str:
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    if isinstance(date_str, date):
        return date_str
    date_str = str(date_str).strip()
    # بررسی فرمت شمسی (سال > 1300)
    if '/' in date_str or '-' in date_str:
        parts = date_str.replace('/', '-').split('-')
        if len(parts) == 3:
            y, m, d = map(int, parts)
            if y > 1300:  # شمسی
                import jdatetime
                return jdatetime.date(y, m, d).togregorian()
            else:  # میلادی
                return date(y, m, d)
    return None

# ---------------------- آپلود فایل و پردازش ----------------------
# ---------------------- آپلود فایل و پردازش پیشرفته ----------------------


def fix_excel_date(value):
    """
    تنظیم تاریخ ورودی از اکسل
    پشتیبانی از: datetime، date، رشته، مقدار خالی
    """
    if value is None:
        return None

    # اگر تاریخ واقعی اکسل باشد
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.date() if isinstance(value, datetime.datetime) else value

    # اگر رشته باشد
    value = str(value).strip()
    if not value:
        return None

    parsed = parse_date(value)
    return parsed




# ---------------------- آپلود و پردازش اکسل ----------------------



# ---------- بهینه‌سازی تابع upload_members_and_payments برای عملیات bulk ----------
@transaction.atomic
@login_required
def upload_members_and_payments(request):
    """
    نسخه بهینه‌شده: خواندن اکسل با openpyxl، تهیه لیست رکوردها در حافظه،
    و سپس انجام bulk_create / bulk_update روی Profile و Payment تا از update_or_create
    در هر حلقه جلوگیری شود.

    نکات:
    - برای ایجاد کاربر جدید از create_user استفاده می‌کنیم تا password هش شود.
    - برای پروفایل‌ها از bulk_create و bulk_update استفاده می‌کنیم.
    - برای واریزی‌ها با یک بار خواندن existing payments از ثبت تکراری جلوگیری می‌کنیم.
    - کل عملیات داخل transaction.atomic اجرا می‌شود تا یا همه اعمال شوند یا هیچ‌کدام.
    """

    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        try:
            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active
        except Exception as e:
            logger.exception("error opening excel file")
            messages.error(request, "❌ خطا در باز کردن فایل. لطفاً فایل اکسل معتبر بارگذاری کنید.")
            return redirect("upload_members_and_payments")

        # خواندن هدرها (فارسی)
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        expected_headers = [
            "نام", "نام خانوادگی", "کد ملی", "شماره همراه", "ایمیل",
            "شماره شناسنامه", "نوبت واریزی", "تاریخ واریز", "مبلغ واریزی"
        ]

        missing_headers = [h for h in expected_headers if h not in headers]
        if missing_headers:
            messages.error(request, f"❌ ستون‌های زیر در فایل وجود ندارند: {', '.join(missing_headers)}")
            return redirect("upload_members_and_payments")

        # جمع‌آوری تمام ردیف‌ها در حافظه (لیست دیکشنری)
        rows = []
        national_codes = set()
        installment_set = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))

            national_code = str(data.get("کد ملی", "")).strip()
            if not national_code:
                continue

            first_name = str(data.get("نام", "")).strip()
            last_name = str(data.get("نام خانوادگی", "")).strip()
            phone = str(data.get("شماره همراه", "")).strip()
            email = str(data.get("ایمیل", "")).strip()
            birth_certificate = str(data.get("شماره شناسنامه", "")).strip()

            installment_number = data.get("نوبت واریزی")
            payment_date = parse_date(data.get("تاریخ واریز"))
            amount = data.get("مبلغ واریزی")

            # نرمال‌سازی مقدارها
            if installment_number is not None:
                try:
                    installment_number = int(installment_number)
                    installment_set.add(installment_number)
                except Exception:
                    installment_number = None

            # جمع‌آوری ردیف
            rows.append({
                "national_code": national_code,
                "first_name": first_name,
                "last_name": last_name,
                "phone": phone,
                "email": email,
                "birth_certificate": birth_certificate,
                "installment_number": installment_number,
                "payment_date": payment_date,
                "amount": amount
            })
            national_codes.add(national_code)

        if not rows:
            messages.warning(request, "هیچ رکورد معتبری در فایل یافت نشد.")
            return redirect("upload_members_and_payments")

        # --------------------------------------------------
        # داده‌های موجود در DB را یک‌بار بخوانیم
        # --------------------------------------------------
        existing_users_qs = User.objects.filter(username__in=national_codes)
        users_dict = {u.username: u for u in existing_users_qs}

        existing_profiles_qs = Profile.objects.filter(user__username__in=national_codes).select_related('user')
        profiles_dict = {p.user.username: p for p in existing_profiles_qs}

        # ApprovedPaymentDate برای شماره اقساط موجود را بگیریم
        approved_dates_qs = ApprovedPaymentDate.objects.filter(installment_number__in=installment_set)
        approved_map = {a.installment_number: a for a in approved_dates_qs}

        # اگر برخی installment در DB نبود، آن‌ها را ایجاد (bulk) کنیم
        missing_installments = [i for i in installment_set if i not in approved_map]
        if missing_installments:
            objs = [ApprovedPaymentDate(installment_number=i, due_date=None) for i in missing_installments]
            ApprovedPaymentDate.objects.bulk_create(objs)
            # refresh map
            for a in ApprovedPaymentDate.objects.filter(installment_number__in=missing_installments):
                approved_map[a.installment_number] = a

        # existing payments set برای جلوگیری از درج تکراری
        existing_payments_qs = Payment.objects.filter(user__username__in=national_codes, installment_number__in=installment_set)
        existing_payments_set = set()
        for p in existing_payments_qs.values_list('user__username', 'installment_number', 'payment_date', 'amount'):
            # مقدار payment_date ممکن است None باشد؛ تبدیل به str برای مقایسه
            existing_payments_set.add((p[0], p[1], str(p[2]) if p[2] is not None else None, float(p[3]) if p[3] is not None else None))

        # --------------------------------------------------
        # آماده‌سازی لیست‌های bulk
        # --------------------------------------------------
        new_users = []         # لیست User (objects) که باید ایجاد شوند
        new_profiles = []      # لیست Profile برای bulk_create
        update_profiles = []   # لیست Profile برای bulk_update
        new_payments = []      # لیست Payment برای bulk_create

        added_users = 0
        added_payments = 0

        # برای هر ردیف در اکسل، اشیاء مربوطه را در لیست‌ها قرار می‌دهیم
        for r in rows:
            nc = r['national_code']

            # --- user ---
            user = users_dict.get(nc)
            if not user:
                # ایجاد کاربر جدید (با هش کردن پسورد)
                user = User.objects.create_user(username=nc, password=nc, first_name=r['first_name'], last_name=r['last_name'], email=r['email'])
                users_dict[nc] = user
                added_users += 1

            # --- profile ---
            profile = profiles_dict.get(nc)
            if profile:
                # به‌روزرسانی در حافظه
                changed = False
                if profile.phone_number != r['phone']:
                    profile.phone_number = r['phone']
                    changed = True
                if profile.birth_certificate != r['birth_certificate']:
                    profile.birth_certificate = r['birth_certificate']
                    changed = True
                if profile.national_code != nc:
                    profile.national_code = nc
                    changed = True
                if changed:
                    update_profiles.append(profile)
            else:
                p = Profile(user=user, phone_number=r['phone'], birth_certificate=r['birth_certificate'], national_code=nc)
                new_profiles.append(p)
                profiles_dict[nc] = p

            # --- payment ---
            inst = r['installment_number']
            pay_dt = r['payment_date']
            amt = r['amount']
            if inst and pay_dt and amt:
                key = (nc, inst, str(pay_dt) if pay_dt is not None else None, float(amt) if amt is not None else None)
                if key in existing_payments_set:
                    continue

                due_date_val = approved_map.get(inst).due_date if approved_map.get(inst) else None

                payment = Payment(
                    user=user,
                    installment_number=inst,
                    amount=amt,
                    payment_date=pay_dt,
                    due_date=due_date_val
                )
                new_payments.append(payment)
                added_payments += 1

        # --------------------------------------------------
        # اجرای عملیات bulk در DB
        # --------------------------------------------------
        if new_profiles:
            Profile.objects.bulk_create(new_profiles, batch_size=500)

        if update_profiles:
            # مشخص کردن فیلدهای قابل به‌روزرسانی
            Profile.objects.bulk_update(update_profiles, ['phone_number', 'birth_certificate', 'national_code'], batch_size=500)

        if new_payments:
            Payment.objects.bulk_create(new_payments, batch_size=500)

        msg = (
            f"✅ فایل با موفقیت پردازش شد. "
            f"{added_users} کاربر جدید، "
            f"{added_payments} رکورد واریزی ثبت شد."
        )
        messages.success(request, msg)
        return redirect("upload_members_and_payments")

    return render(request, "accounts/upload_members_and_payments.html")






# ---------------------- ورود کاربر ----------------------
from django.contrib.auth.views import LoginView
from django.contrib import messages

class CustomLoginView(LoginView):
    def get_template_names(self):
        return [self.template_name]

    def get_success_url(self):
        user = self.request.user

        # ۱. اگر سوپریوزر است (ادمین اصلی)
        if user.is_superuser:
            return '/admin/'

        # ۲. اگر در گروه مدیران آپلود است
        elif user.groups.filter(name='upload_manager').exists():
            return '/upload-members/'

        # ۳. در غیر این صورت، کاربر معمولی است
        else:
            return '/dashboard/'

    def form_valid(self, form):
        messages.success(self.request, f"خوش آمدید {self.request.user.username} 🌷")
        return super().form_valid(form)


#-------------------------------------------------------
def member_score_view(request):
    users = User.objects.all().prefetch_related('payment_set')
    
    scores = []

    for u in users:
        total_score = 0
        today = date.today()
        for p in u.payment_set.all():
            if p.payment_date and p.due_date:
                if p.payment_date > p.due_date:
                    diff_days = (today - p.payment_date).days
                else:
                    diff_days = (today - p.due_date).days
            else:
                diff_days = 0

            total_score += (diff_days * float(p.amount)) / 100_000_000

        scores.append((u.id, total_score))

    # سورت نزولی
    scores.sort(key=lambda x: x[1], reverse=True)

    # پیدا کردن رتبه
    rank = 1
    user_rank = None
    for uid, score in scores:
        if uid == request.user.id:
            user_rank = rank
            break
        rank += 1

    # محاسبه امتیاز کاربر جاری
    member_total, _ = calculate_total_score(request.user)

    context = {
        "member": request.user,
        "total_score": member_total,
        "rank": user_rank,
    }

    return render(request, "member_score.html", context)

def logout_view(request):
    logout(request)
    return redirect('login')  # بعد از خروج، به صفحه ورود برگردد
