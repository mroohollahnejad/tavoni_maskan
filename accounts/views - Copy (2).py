from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib.auth import login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponse
from django.db import transaction, IntegrityError
from django.views.decorators.csrf import csrf_exempt
from .models import Profile, Payment, ApprovedPaymentDate
from .forms import ProfileForm, UserUpdateForm
import jdatetime
from datetime import datetime, date
import openpyxl
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)

# ================== متغیر سراسری عناوین ستون‌های اکسل ==================
EXCEL_HEADERS = [
    "نام", "نام خانوادگی", "کد ملی", "شماره همراه", "ایمیل",
    "شماره شناسنامه", "نوبت واریزی", "تاریخ واریز", "مبلغ واریزی"
]

# ---------------------- تبدیل تاریخ ----------------------
def parse_date(date_str):
    if not date_str:
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    if isinstance(date_str, date):
        return date_str
    date_str = str(date_str).strip()
    if '/' in date_str or '-' in date_str:
        parts = date_str.replace('/', '-').split('-')
        if len(parts) == 3:
            y, m, d = map(int, parts)
            if y > 1300:
                return jdatetime.date(y, m, d).togregorian()
            else:
                return date(y, m, d)
    return None

def to_jalali(date_obj):
    if not date_obj:
        return ''
    if isinstance(date_obj, jdatetime.date):
        return date_obj.strftime("%Y-%m-%d")
    return jdatetime.date.fromgregorian(date=date_obj).strftime("%Y-%m-%d")

# ================== تمیز کردن کدملی ==================
def clean_national_code(raw_nc):
    if not raw_nc:
        return ""
    s = str(raw_nc).strip()
    s = s.strip("'\"")
    s = ''.join(filter(str.isdigit, s))
    if not s:
        return ""
    s = s.lstrip('0')
    if not s:
        s = "0"
    return s[-10:].zfill(10)

# ================== پروفایل امن ==================
def get_or_create_profile_safe(user, phone, bc, nc):
    try:
        profile, created = Profile.objects.get_or_create(
            user=user,
            defaults={
                "phone_number": phone,
                "birth_certificate": bc,
                "national_code": nc
            }
        )
        updated = False
        if profile.phone_number != phone:
            profile.phone_number = phone
            updated = True
        if profile.birth_certificate != bc:
            profile.birth_certificate = bc
            updated = True
        if profile.national_code != nc:
            profile.national_code = nc
            updated = True
        if updated:
            profile.save()
        return profile
    except IntegrityError:
        return Profile.objects.get(user=user)

# ---------------------- محاسبه امتیاز کل ----------------------
def calculate_total_score(user):
    payments = Payment.objects.filter(user=user).order_by('installment_number')
    total_score = 0
    payments_data = []
    today = date.today()
    for p in payments:
        if p.payment_date and p.due_date:
            if p.payment_date > p.due_date:
                diff_days = (today - p.payment_date).days
            else:
                diff_days = (today - p.due_date).days
        else:
            diff_days = 0
        score = (diff_days * int(p.amount)) / 100_000_000
        total_score += score
        payments_data.append({
            'id': p.id,
            'installment_number': p.installment_number,
            'amount': f"{int(p.amount):,}",
            'payment_date': p.payment_date,
            'payment_date_j': to_jalali(p.payment_date),
            'due_date': p.due_date,
            'due_date_j': to_jalali(p.due_date),
            'diff_days': diff_days,
            'score': round(score, 2)
        })
    return total_score, payments_data

# ================== تولید و دانلود فایل نمونه ==================
def generate_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(EXCEL_HEADERS)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    sample_row = ["محمد", "روح‌الله‌نژاد", "۳۳۹۲۰۴۸۴۴۴","09121234567", "test@example.com",
                  "123456", 1, "1404-01-15", 1000000]
    ws.append(sample_row)
    return wb

@login_required
def download_sample_excel(request):
    wb = generate_sample_excel()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_members.xlsx"'
    wb.save(response)
    return response

# ================== آپلود و پردازش اکسل ==================
@transaction.atomic
@login_required
def upload_members_and_payments(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active
        except:
            messages.error(request, "❌ خطا در باز کردن فایل. لطفاً فایل اکسل معتبر بارگذاری کنید.")
            return redirect("upload_members_and_payments")

        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        missing_headers = [h for h in EXCEL_HEADERS if h not in headers]
        if missing_headers:
            messages.error(request, f"❌ ستون‌های زیر در فایل وجود ندارند: {', '.join(missing_headers)}")
            return redirect("upload_members_and_payments")

        rows = []
        national_codes = set()
        installment_set = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            raw_nc = data.get("کد ملی")
            nc = clean_national_code(raw_nc)
            if not nc or len(nc) != 10:
                messages.warning(request, f"ردیف نادیده گرفته شد: کدملی نامعتبر → {raw_nc}")
                continue

            first = str(data.get("نام") or "").strip()
            last = str(data.get("نام خانوادگی") or "").strip()
            phone = str(data.get("شماره همراه") or "").strip()
            email = str(data.get("ایمیل") or "").strip()
            bc = str(data.get("شماره شناسنامه") or "").strip()
            inst = data.get("نوبت واریزی")
            pay_dt = parse_date(data.get("تاریخ واریز"))
            amt = data.get("مبلغ واریزی")
            try:
                inst = int(inst)
                installment_set.add(inst)
            except:
                inst = None

            rows.append({
                "national_code": nc,
                "first_name": first,
                "last_name": last,
                "phone": phone,
                "email": email,
                "birth_certificate": bc,
                "installment_number": inst,
                "payment_date": pay_dt,
                "amount": amt,
            })
            national_codes.add(nc)

        if not rows:
            messages.warning(request, "هیچ رکورد معتبری در فایل یافت نشد.")
            return redirect("upload_members_and_payments")

        users_dict = {u.username: u for u in User.objects.filter(username__in=national_codes)}
        approved_map = {
            a.installment_number: a
            for a in ApprovedPaymentDate.objects.filter(installment_number__in=installment_set)
        }
        missing_inst = [i for i in installment_set if i not in approved_map]
        if missing_inst:
            ApprovedPaymentDate.objects.bulk_create(
                [ApprovedPaymentDate(installment_number=i, due_date=None) for i in missing_inst]
            )
            for a in ApprovedPaymentDate.objects.filter(installment_number__in=missing_inst):
                approved_map[a.installment_number] = a

        existing_payments_set = set(
            Payment.objects.filter(
                user__username__in=national_codes,
                installment_number__in=installment_set
            ).values_list("user__username", "installment_number", "payment_date", "amount")
        )

        new_payments = []
        added_users = 0
        added_payments = 0
        for r in rows:
            nc = r["national_code"]
            user = users_dict.get(nc)
            if not user:
                user = User.objects.create_user(
                    username=nc,
                    password=nc,
                    first_name=r["first_name"],
                    last_name=r["last_name"],
                    email=r["email"],
                )
                users_dict[nc] = user
                added_users += 1

            profile = get_or_create_profile_safe(
                user=user,
                phone=r["phone"],
                bc=r["birth_certificate"],
                nc=nc
            )

            inst = r["installment_number"]
            pay_dt = r["payment_date"]
            amt = r["amount"]

            if inst and pay_dt and amt:
                key = (nc, inst, pay_dt, amt)
                if key not in existing_payments_set:
                    due_val = approved_map.get(inst).due_date if approved_map.get(inst) else None
                    new_payments.append(
                        Payment(
                            user=user,
                            installment_number=inst,
                            payment_date=pay_dt,
                            amount=amt,
                            due_date=due_val,
                        )
                    )
                    added_payments += 1

        if new_payments:
            Payment.objects.bulk_create(new_payments, batch_size=500)

        messages.success(
            request,
            f"✅ پردازش موفق: {added_users} کاربر جدید، {added_payments} واریزی جدید ثبت شد."
        )
        return redirect("upload_members_and_payments")

    return render(request, "accounts/upload_members_and_payments.html")

# ---------------------- داشبورد ----------------------
@login_required
def dashboard(request):
    user = request.user
    profile = Profile.objects.get_or_create(user=user)[0]
    payments = Payment.objects.filter(user=user).order_by('installment_number')
    approved_dates = ApprovedPaymentDate.objects.all()
    total_score, payments_data = calculate_total_score(user)
    approved_dates_list = {p.installment_number: p.due_date for p in approved_dates}

    users = User.objects.all()
    scores = []
    for u in users:
        score, _ = calculate_total_score(u)
        scores.append((u.id, score))
    scores.sort(key=lambda x: x[1], reverse=True)
    user_rank = None
    rank = 1
    for uid, score in scores:
        if uid == user.id:
            user_rank = rank
            break
        rank += 1

    context = {
        'user': user,
        'profile': profile,
        'payments': payments_data,
        'approved_dates': approved_dates_list,
        'installments': [d.installment_number for d in ApprovedPaymentDate.objects.all().order_by('installment_number')],
        'active_tab': request.GET.get('tab', 'personal'),
        'total_score': total_score,
        "rank": user_rank
    }
    return render(request, 'accounts/dashboard.html', context)

# ================== AJAX واریزی ==================
@csrf_exempt
@login_required
def payment_create_ajax(request):
    if request.method == 'POST':
        try:
            installment = int(request.POST.get('installment_number'))
            amount = int(float(request.POST.get('amount', 0)))
            payment_date = parse_date(request.POST.get('payment_date'))
            due_date_obj = ApprovedPaymentDate.objects.filter(installment_number=installment).first()
            due_date_val = due_date_obj.due_date if due_date_obj else None
            payment = Payment.objects.create(
                user=request.user,
                installment_number=installment,
                amount=amount,
                payment_date=payment_date,
                due_date=due_date_val
            )
            total_score, _ = calculate_total_score(request.user)
            return JsonResponse({
                'status': 'success',
                'message': '✅ واریزی ثبت شد',
                'payment': {
                    'id': payment.id,
                    'installment_number': payment.installment_number,
                    'amount': f"{payment.amount:,}",
                    'payment_date': to_jalali(payment.payment_date),
                    'due_date': to_jalali(payment.due_date)
                },
                'total_score': total_score
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر ❌'})

@csrf_exempt
@login_required
def payment_edit_ajax(request, pk):
    if request.method == 'POST':
        try:
            payment = get_object_or_404(Payment, pk=pk, user=request.user)
            field = request.POST.get('field')
            value = request.POST.get('value')
            if field == 'installment_number':
                payment.installment_number = int(value)
                due_date_obj = ApprovedPaymentDate.objects.filter(installment_number=value).first()
                payment.due_date = due_date_obj.due_date if due_date_obj else None
            elif field == 'amount':
                payment.amount = int(float(value))
            elif field == 'payment_date':
                payment.payment_date = parse_date(value)
            payment.save()
            total_score, _ = calculate_total_score(request.user)
            return JsonResponse({
                'status': 'success',
                'message': '✅ ویرایش انجام شد',
                'amount': f"{payment.amount:,}",
                'payment_date': to_jalali(payment.payment_date),
                'due_date': to_jalali(payment.due_date),
                'total_score': total_score
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر ❌'})

@csrf_exempt
@login_required
def payment_delete_ajax(request, pk):
    if request.method == 'POST':
        payment = get_object_or_404(Payment, pk=pk, user=request.user)
        payment.delete()
        total_score, _ = calculate_total_score(request.user)
        return JsonResponse({'status': 'success', 'message': '🗑️ واریزی حذف شد', 'total_score': total_score})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر ❌'})

@login_required
def score_ajax(request):
    total_score, payments_data = calculate_total_score(request.user)
    payments_json = [
        {
            'installment_number': p['installment_number'],
            'amount': int(p['amount'].replace(",", "")),
            'payment_date': p['payment_date_j'],
            'due_date': p['due_date_j'],
            'diff_days': p['diff_days'],
            'score': p['score']
        } for p in payments_data
    ]
    return JsonResponse({'total_score': total_score, 'payments': payments_json})

# ================== پروفایل ==================
@login_required
def profile_edit(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, "✅ تغییرات ذخیره شد")
            return redirect('dashboard')
        else:
            messages.error(request, "❌ خطا در فرم")
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileForm(instance=profile)
    return render(request, 'accounts/profile_edit.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

# ================== ثبت‌نام ==================
def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        if password != password2:
            messages.error(request, 'رمزها مطابقت ندارند ❌')
            return redirect('register')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'این نام کاربری قبلاً ثبت شده است ❌')
            return redirect('register')
        user = User.objects.create_user(username=username, password=password)
        login(request, user)
        messages.success(request, 'ثبت‌نام با موفقیت انجام شد ✅')
        return redirect('dashboard')
    return render(request, 'accounts/register.html')

# ================== ورود ==================
class CustomLoginView(LoginView):
    def get_template_names(self):
        return [self.template_name]
    def get_success_url(self):
        user = self.request.user
        if user.is_superuser:
            return '/admin/'
        elif user.groups.filter(name='upload_manager').exists():
            return '/upload-members/'
        else:
            return '/dashboard/'
    def form_valid(self, form):
        messages.success(self.request, f"خوش آمدید {self.request.user.username} 🌷")
        return super().form_valid(form)

# ================== خروج ==================
def logout_view(request):
    logout(request)
    return redirect('login')

# ================== امتیاز کاربر ==================
@login_required
def member_score_view(request):
    users = User.objects.all().prefetch_related('payment_set')
    scores = []
    for u in users:
        total_score = 0
        today = date.today()
        for p in u.payment_set.all():
            if p.payment_date and p.due_date:
                if p.payment_date > p.due_date:
                    diff_days = (today - p.payment_date).days
                else:
                    diff_days = (today - p.due_date).days
            else:
                diff_days = 0
            total_score += (diff_days * int(p.amount)) / 100_000_000
        scores.append({'user': u, 'score': total_score})
    scores.sort(key=lambda x: x['score'], reverse=True)
    for idx, s in enumerate(scores, 1):
        s['rank'] = idx
    return render(request, 'accounts/member_scores.html', {'scores': scores})
